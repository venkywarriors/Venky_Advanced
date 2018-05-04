'''
Created on Nov 24, 2017

@author: venkateshwara.d

'''

import os
import openpyxl


dir_path = os.path.dirname(os.path.realpath(__file__))

Excel_path = dir_path+"\TestData\Testdata.xlsx"

def createFolder(directory):   # './Testdata/'
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
            print('Folder created successfully')
        else:
            print('Folder already exists')                
    except OSError:
        print ('Error: Creating directory. ' +  directory)
        
def writeData(cellvalue,testdata,sheetname):  
    try:
        if os.path.isfile(Excel_path)== True :
            print("opens "+Excel_path)
            excel_document = openpyxl.load_workbook(Excel_path)          
            try: #KeyError:
                sheet = excel_document.get_sheet_by_name(str(sheetname))
            except KeyError as e:    
                print('Worksheet'+ str(sheetname) +' does not exist '+e.value)
            sheet[str(cellvalue)] = str(testdata) 
            print("cell value "+cellvalue+" == "+testdata) 
            excel_document.save(Excel_path)         
    except FileNotFoundError as e:
        print('FileNotFoundError , value:', e.value)  
    except Exception: 
        print('Permission denied close the file and then execute')  
                  
        
def readDataRowAndColumns(column,row,sheetname):
    try:
        if os.path.isfile(Excel_path)== True :
            print("opens "+Excel_path)
            excel_document = openpyxl.load_workbook(Excel_path)
            try: #KeyError:
                sheet = excel_document.get_sheet_by_name(str(sheetname))
            except KeyError as e:    
                print('Worksheet'+ str(sheetname) +' does not exist '+e.value)
            return sheet.cell(row=int(row),column=int(column)).value
    except FileNotFoundError as e:
        print('FileNotFoundError , value:', e.value)
    except Exception as e:
        print( e.value)     
            
def writeDataRowAndColumns(column,row,testdata,sheetname):  
    try:
        if os.path.isfile(Excel_path)== True :
            print("opens "+Excel_path)
            excel_document = openpyxl.load_workbook(Excel_path)          
            try: #KeyError:
                sheet = excel_document.get_sheet_by_name(str(sheetname))
            except KeyError as e:    
                print('Worksheet'+ str(sheetname) +' does not exist '+e.value)
            sheet.cell(row=int(row),column=int(column)).value = str(testdata) 
            print("cell value row "+str(row)+" column = "+str(column)+" == "+testdata) 
            excel_document.save(Excel_path)         
    except FileNotFoundError as e:
        print('FileNotFoundError , value:', e.value)  
    except Exception as e:
        print('Permission denied close the file and then execute', e.value)          
        
def readData(cellvalue,sheetname):  
    try:
        if os.path.isfile(Excel_path)== True :
            print("opens "+Excel_path)
            excel_document = openpyxl.load_workbook(Excel_path)
            try: #KeyError:
                sheet = excel_document.get_sheet_by_name(str(sheetname))
            except KeyError as e:    
                print('Worksheet'+ str(sheetname) +' does not exist '+e.value)
            return sheet[str(cellvalue)].value
    except FileNotFoundError as e:
        print('FileNotFoundError , value:', e.value)
        
a=readData("G3", "HG") 
print('output of g3',a)  
b=readDataRowAndColumns(9, 8, "HG")  
print('output of 7,8',b)  
writeData("G18", "venkateshwara D", "HG")  
writeDataRowAndColumns(6,19,"venkateshwara D", "Sheet3")   
        
       
     
    

